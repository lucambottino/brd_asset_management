import pandas as pd
import numpy as np
import re
from decimal import Decimal, ROUND_HALF_UP

# ============================================================
# Helpers
# ============================================================

def brl_to_decimal(x) -> Decimal:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return Decimal("0")
    if isinstance(x, Decimal):
        return x
    if isinstance(x, (int, float, np.integer, np.floating)):
        return Decimal(str(x))
    s = str(x).strip()
    s = s.replace("R$", "").replace("\u00a0", " ").strip()
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return Decimal(s)

def pct_to_decimal(x) -> Decimal:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return Decimal("0")
    if isinstance(x, Decimal):
        return x
    if isinstance(x, (int, float, np.integer, np.floating)):
        return Decimal(str(x)) / Decimal("100")
    s = str(x).strip().replace("%", "").replace("\u00a0", " ")
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    return Decimal(s) / Decimal("100")

def q2(x: Decimal) -> Decimal:
    return x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def fmt_brl(d: Decimal) -> str:
    d = q2(d)
    s = f"{d:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"

def fmt_pct(frac: Decimal) -> str:
    p = q2(frac * Decimal("100"))
    s = f"{p:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}%"

def allocate_with_residual(total: Decimal, weights: pd.Series) -> pd.Series:
    if total == 0:
        return pd.Series([Decimal("0.00")] * len(weights), index=weights.index)

    w = weights.astype("float64").values
    if np.allclose(w.sum(), 0.0):
        w = np.ones_like(w, dtype="float64")

    raw = [Decimal(str(total)) * Decimal(str(wi)) / Decimal(str(w.sum())) for wi in w]
    rounded = [q2(a) for a in raw]
    allocated = pd.Series(rounded, index=weights.index, dtype=object)

    residual = total - sum(allocated.tolist(), start=Decimal("0.00"))
    cent = Decimal("0.01")
    steps = int((residual / cent).to_integral_value(rounding=ROUND_HALF_UP))

    if steps != 0:
        remainders = pd.Series([r - q2(r) for r in raw], index=weights.index, dtype=object)
        if steps > 0:
            order = remainders.sort_values(ascending=False).index.tolist()
            for i in range(steps):
                allocated.loc[order[i % len(order)]] = allocated.loc[order[i % len(order)]] + cent
        else:
            order = remainders.sort_values(ascending=True).index.tolist()
            for i in range(abs(steps)):
                allocated.loc[order[i % len(order)]] = allocated.loc[order[i % len(order)]] - cent

    final_sum = sum(allocated.tolist(), start=Decimal("0.00"))
    if final_sum != total:
        allocated.iloc[0] = allocated.iloc[0] + (total - final_sum)

    return allocated

def sum_decimals_str(series: pd.Series) -> str:
    return str(sum((Decimal(x) for x in series), start=Decimal("0.00")))

# ============================================================
# Core computation (extended output)
# ============================================================

def compute_yield_by_account(
    lft_lots_df: pd.DataFrame,
    accounts_df: pd.DataFrame,
    *,
    weight_mode: str = "quantity",  # "quantity" or "percentage"
) -> pd.DataFrame:
    lft = lft_lots_df.copy()
    acc = accounts_df.copy()

    # Dates
    lft["Data de compra"] = pd.to_datetime(lft["Data de compra"], dayfirst=True)
    acc["Data de Compra"] = pd.to_datetime(acc["Data de Compra"], dayfirst=True)

    # Money fields
    money_cols = ["Valor de compra", "Valor bruto", "Valor líquido", "IOF", "IR"]
    for col in money_cols:
        if col not in lft.columns:
            raise ValueError(f"Missing column in lft_lots_df: {col}")
        lft[col] = lft[col].apply(brl_to_decimal)

    # Derived lot fields
    lft["Rendimento bruto"] = lft["Valor bruto"] - lft["Valor de compra"]
    lft["Rendimento líquido"] = lft["Valor líquido"] - lft["Valor de compra"]

    # Aggregate lots by date
    lots_by_date = (
        lft.groupby("Data de compra", as_index=False)
           .agg(
               Investido=("Valor de compra", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
               Valor_Bruto=("Valor bruto", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
               Valor_Liquido=("Valor líquido", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
               IOF=("IOF", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
               IR=("IR", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
               Rendimento_Bruto=("Rendimento bruto", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
               Rendimento_Liquido=("Rendimento líquido", lambda s: sum(s.tolist(), start=Decimal("0.00"))),
           )
    )

    # Weights
    if weight_mode not in ("quantity", "percentage"):
        raise ValueError("weight_mode must be 'quantity' or 'percentage'")

    if weight_mode == "quantity":
        if "Quantity" not in acc.columns:
            raise ValueError("Missing column in accounts_df: Quantity")
        acc["_w"] = acc["Quantity"].astype("int64")
    else:
        if "Percentage" not in acc.columns:
            raise ValueError("Missing column in accounts_df: Percentage")
        acc["_w"] = acc["Percentage"].apply(pct_to_decimal).apply(lambda d: float(d))

    out_rows = []

    for _, lot in lots_by_date.iterrows():
        dt = lot["Data de compra"]
        subset = acc.loc[acc["Data de Compra"] == dt].copy()
        if subset.empty:
            # if you want, you can log/print here
            continue

        alloc_invest = allocate_with_residual(lot["Investido"], subset["_w"])
        alloc_iof    = allocate_with_residual(lot["IOF"], subset["_w"])
        alloc_ir     = allocate_with_residual(lot["IR"], subset["_w"])
        alloc_gross  = allocate_with_residual(lot["Rendimento_Bruto"], subset["_w"])
        alloc_net    = allocate_with_residual(lot["Rendimento_Liquido"], subset["_w"])

        subset["Investido (BRL)"] = alloc_invest.values
        subset["IOF (BRL)"] = alloc_iof.values
        subset["IR (BRL)"] = alloc_ir.values
        subset["Rendimento Bruto (BRL)"] = alloc_gross.values
        subset["Rendimento Líquido (BRL)"] = alloc_net.values

        def safe_div(num: Decimal, den: Decimal) -> Decimal:
            if den == 0:
                return Decimal("0")
            return num / den

        subset["Lucro Bruto %"] = subset.apply(
            lambda r: safe_div(r["Rendimento Bruto (BRL)"], r["Investido (BRL)"]),
            axis=1,
        )
        subset["Lucro Líquido %"] = subset.apply(
            lambda r: safe_div(r["Rendimento Líquido (BRL)"], r["Investido (BRL)"]),
            axis=1,
        )

        for _, r in subset.iterrows():
            out_rows.append(
                {
                    "Account": r["Account"],
                    "Token Account": r.get("Token Account", None),
                    "Data de Compra": r["Data de Compra"].date().isoformat(),
                    "Weight Mode": weight_mode,
                    "Weight": r["_w"],
                    "Quantidade (token)": int(r["Quantity"]) if "Quantity" in subset.columns else None,
                    "Investido (BRL)": str(r["Investido (BRL)"]),
                    "IOF (BRL)": str(r["IOF (BRL)"]),
                    "IR (BRL)": str(r["IR (BRL)"]),
                    "Rendimento Bruto (BRL)": str(r["Rendimento Bruto (BRL)"]),
                    "Rendimento Líquido (BRL)": str(r["Rendimento Líquido (BRL)"]),
                    "Lucro Bruto %": str(r["Lucro Bruto %"]),
                    "Lucro Líquido %": str(r["Lucro Líquido %"]),
                }
            )

    out = pd.DataFrame(out_rows)
    if out.empty:
        return out

    # Totals per account across dates
    for col in ["Investido (BRL)", "IOF (BRL)", "IR (BRL)", "Rendimento Bruto (BRL)", "Rendimento Líquido (BRL)"]:
        totals = (
            out.groupby("Account", as_index=False)[col]
               .apply(sum_decimals_str)
               .rename(columns={col: f"{col} Total"})
        )
        out = out.merge(totals, on="Account", how="left")

    # Total % from totals
    def to_dec(s): return Decimal(s)

    out["Lucro Bruto % Total"] = out.apply(
        lambda r: str((to_dec(r["Rendimento Bruto (BRL) Total"]) / to_dec(r["Investido (BRL) Total"]))
                      if to_dec(r["Investido (BRL) Total"]) != 0 else Decimal("0")),
        axis=1,
    )
    out["Lucro Líquido % Total"] = out.apply(
        lambda r: str((to_dec(r["Rendimento Líquido (BRL) Total"]) / to_dec(r["Investido (BRL) Total"]))
                      if to_dec(r["Investido (BRL) Total"]) != 0 else Decimal("0")),
        axis=1,
    )

    return out

# ============================================================
# DEFINE YOUR INPUT TABLES HERE (this fixes NameError)
# ============================================================

lft_lots = pd.DataFrame(
    [
        {
            "Data de compra": "23/01/2026",
            "Valor de compra": "R$ 419.656,79",
            "Valor líquido": "R$ 423712,36",
            "Valor bruto": "R$ 424.889,78",
            "IOF": "R$ 0",
            "IR": "R$ 1177,42",
        },
        {
            "Data de compra": "30/01/2026",
            "Valor de compra": "R$ 256.148,18",
            "Valor líquido": "R$ 257955,14",
            "Valor bruto": "R$ 258.628,56",
            "IOF": "R$ 148,82",
            "IR": "R$ 524,60",
        },
    ]
)

accounts = pd.DataFrame(
    [
        {"Account": "AiYxij88KRXxWiAhNMqzxBd33wu5CSUd6KPAdoZETgfL", "Token Account": "5MZZn4D96MkndVoBjSpuqCdnMDuv87VFV1A6kZrHLuA4", "Quantity": 260870, "Percentage": "38.24", "Data de Compra": "30/01/2026"},
        {"Account": "6pgtAbfCwukveUfKYm22FEHP8w7NKD9Vd73NaCExLAKi", "Token Account": "58kQw1fVjtSdKHYHBsvRenDPfDFSV8MvwesVWsE8MsZM", "Quantity": 60000, "Percentage": "8.80", "Data de Compra": "23/01/2026"},
        {"Account": "CbVUNS2kgnJszuWwP7ioxGLkUHGWppbcAitC5AynnDSx", "Token Account": "4f2BJC3B3LRiapnz8xiT4YkXLvQzHbnCFxbMaSito4fh", "Quantity": 60000, "Percentage": "8.80", "Data de Compra": "23/01/2026"},
        {"Account": "czTNJnh5ihaYfaWzSjjratZrqMcw318KAPBQHdsvVBf", "Token Account": "HnRpBvc3gMfSrJzCer36jnobnHWTph56hFKyvKG91iYP", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "Ai6gzjwzCL9E1rVMR9gP5gD9Y9Eds1uKj1NoC3ggJw4T", "Token Account": "7Sxy6aovd4VEThL7P8X6ZcS9DxbH9Sy6HkMMSeisVf8b", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "9bxQUSioYmCfqeraQ4x1UND5k2PLMmMxLDMCGxVW9Mzz", "Token Account": "4bNa2pBXGVra7kqixKsGYkwx3MdSCuPiZpGrC7Tsc3Lf", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "8wju6VrZpjoZC5TZTGWVJSjNMnz8qeKweZpArcu8BoX9", "Token Account": "2CvmUYrv4CcZR2zvJhttNMLy4zNKKicwWoBRzBEveXx6", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "3qsNsvWUJsZTk9FS3mmFv2YiUD6gPUJwpeqLQv4ahof8", "Token Account": "2WvvG3sHfHyRBajuzCxbREedRHNGX1YxiEVohGK5GEEf", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "3hJgnmP5gZfkV876PnvJpzCUrq9vGm9fckQNhUB6ftWV", "Token Account": "FPzm4TruXnfGkf2t3YgC5vCPjkLxwYGvPCJXTc4UP1WC", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "5h1sZWC93q9mWihsa2tvrsuJoqGY58655CSHkug2opRV", "Token Account": "FMb9i4T61D6YNh8cBWvMYKVuPMV6F1ETq2ZV9GLJU51e", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "GMuR2i1TfXKvdXJn3iBZCdXomnqVQN3NpEApnmixEndb", "Token Account": "2Sm5kBgNey4zMnxaGBxNN8HhsoikbmgwEJu4cy9jfrS1", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "3DHW2ZCYcKdzqBxyeVcG6TxG6foHdokvJS6Xr55VKPZq", "Token Account": "CSXsNbG9jGk4hYwVz1W2DTXX4TX7dRZ9bfypHgxtuDfV", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "5C7R3fhGJ37wBWEbSrw8KU8sfj8zeoctDAFDkLxcJmgv", "Token Account": "5cKSj1EmK8tViMjKGnaKWSD2jBFK5FN6aQ8hEPf3vTor", "Quantity": 30000, "Percentage": "4.40", "Data de Compra": "23/01/2026"},
        {"Account": "GuU4YH1v6DdkbZwh5Qi7prDxEupGFTtUaTU7EpzRHbQU", "Token Account": "HV2LwjQJj7PCFJGYRWVPZN14YPfEEaTKy4YanFaDJrMx", "Quantity": 1250, "Percentage": "0.18", "Data de Compra": "23/01/2026"},
    ]
)

# ============================================================
# Run
# ============================================================

result = compute_yield_by_account(lft_lots, accounts, weight_mode="quantity")



if result.empty:
    print("No allocations produced. Check date formats and that accounts dates match LFT purchase dates.")
else:
    view_cols = [
        "Account",
        "Data de Compra",
        "Quantidade (token)",
        "Investido (BRL)",
        "IOF (BRL)",
        "IR (BRL)",
        "Rendimento Bruto (BRL)",
        "Rendimento Líquido (BRL)",
        "Lucro Bruto %",
        "Lucro Líquido %",
        "Investido (BRL) Total",
        "IOF (BRL) Total",
        "IR (BRL) Total",
        "Rendimento Bruto (BRL) Total",
        "Rendimento Líquido (BRL) Total",
        "Lucro Bruto % Total",
        "Lucro Líquido % Total",
    ]
    view = result[view_cols].copy()

    # Money formatting
    for c in [
        "Investido (BRL)",
        "IOF (BRL)",
        "IR (BRL)",
        "Rendimento Bruto (BRL)",
        "Rendimento Líquido (BRL)",
        "Investido (BRL) Total",
        "IOF (BRL) Total",
        "IR (BRL) Total",
        "Rendimento Bruto (BRL) Total",
        "Rendimento Líquido (BRL) Total",
    ]:
        view[c] = view[c].apply(lambda s: fmt_brl(Decimal(s)))

    # Percent formatting
    for c in ["Lucro Bruto %", "Lucro Líquido %", "Lucro Bruto % Total", "Lucro Líquido % Total"]:
        view[c] = view[c].apply(lambda s: fmt_pct(Decimal(s)))

    print(view.to_string(index=False))


# result_formatted = result[["Account", "Data de Compra", ]]

# print(result)
# result.to_excel("yield_by_account.xlsx", index=False)

import pandas as pd
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def save_result_to_excel(result: pd.DataFrame, out_path: str = "rateio_lft.xlsx") -> str:
    cols = [
        "Account","Data de Compra","Investido (BRL)","IOF (BRL)","IR (BRL)",
        "Rendimento Bruto (BRL)","Rendimento Líquido (BRL)","Lucro Bruto %","Lucro Líquido %"
    ]
    missing = [c for c in cols if c not in result.columns]
    if missing:
        raise ValueError(f"Missing columns in result: {missing}")

    df = result[cols].copy()

    # Convert types for proper Excel number formatting
    df["Data de Compra"] = pd.to_datetime(df["Data de Compra"], errors="coerce")

    money_cols = ["Investido (BRL)","IOF (BRL)","IR (BRL)","Rendimento Bruto (BRL)","Rendimento Líquido (BRL)"]
    pct_cols   = ["Lucro Bruto %","Lucro Líquido %"]

    # Our pipeline stores decimals as strings; convert robustly
    for c in money_cols:
        df[c] = df[c].astype(str).map(lambda s: float(Decimal(s)))
    for c in pct_cols:
        df[c] = df[c].astype(str).map(lambda s: float(Decimal(s)))  # already fraction (e.g., 0.0057)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rateio LFT"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write header
    ws.append(cols)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for _, r in df.iterrows():
        ws.append([
            r["Account"],
            r["Data de Compra"].to_pydatetime() if pd.notna(r["Data de Compra"]) else None,
            r["Investido (BRL)"],
            r["IOF (BRL)"],
            r["IR (BRL)"],
            r["Rendimento Bruto (BRL)"],
            r["Rendimento Líquido (BRL)"],
            r["Lucro Bruto %"],
            r["Lucro Líquido %"],
        ])

    # Number formats
    brl_fmt = u'R$ #,##0.00'
    pct_fmt = '0.00%'

    date_col = cols.index("Data de Compra") + 1
    money_idx = [cols.index(c) + 1 for c in money_cols]
    pct_idx   = [cols.index(c) + 1 for c in pct_cols]

    # Apply formats + alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        # date
        row[date_col - 1].number_format = "dd/mm/yyyy"
        # money
        for j in money_idx:
            row[j - 1].number_format = brl_fmt
            row[j - 1].alignment = Alignment(horizontal="right", vertical="center")
        # percent
        for j in pct_idx:
            row[j - 1].number_format = pct_fmt
            row[j - 1].alignment = Alignment(horizontal="right", vertical="center")
        # account left
        row[0].alignment = Alignment(horizontal="left", vertical="center")

    # Freeze header row and enable filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # Set column widths (nice defaults)
    widths = {
        "Account": 46,
        "Data de Compra": 14,
        "Investido (BRL)": 18,
        "IOF (BRL)": 14,
        "IR (BRL)": 12,
        "Rendimento Bruto (BRL)": 22,
        "Rendimento Líquido (BRL)": 24,
        "Lucro Bruto %": 14,
        "Lucro Líquido %": 16,
    }
    for j, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 16)

    # Turn the range into an Excel "Table" (striped rows, filter dropdowns)
    table_ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    tab = Table(displayName="RateioLFT", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    wb.save(out_path)
    return out_path


# --- usage (after you computed `result`) ---
xlsx_path = save_result_to_excel(result, "rateio_lft.xlsx")
print(f"Saved: {xlsx_path}")
